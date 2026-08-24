"""
Regression tests for convert.py.

Run with:  pytest tests/test_convert.py -v

These tests use a mocked Retail Prices API (`fake_api`) so they run fully
offline and deterministically. Each test corresponds to a specific bug found
during the production review - see the docstring of each test for the bug it
guards against.
"""
import re
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import convert as C


# ── Shared fake Retail Prices API ───────────────────────────────────────────
FX = {"USD": 1.0, "INR": 83.0, "EUR": 0.92, "GBP": 0.79, "AUD": 1.52}
LINUX_HR = {
    "Standard_D4s_v3": 0.192, "Standard_D4_v3": 0.192,
    "Standard_E8s_v5": 0.504,
    "Standard_D4ds_v5": 0.230, "Standard_D4d_v5": 0.230,
    "Standard_M32ms_v2": 3.40,
    "Standard_D2s_v3": 0.096,
    "Standard_D2_v3": 0.096,
}
WIN_DELTA_HR = 0.10
RI1_DISCOUNT = 0.80
RI3_DISCOUNT = 0.60
RHEL_HR = {"Standard_E8s_v5": 0.09}  # only this SKU has a distinct live RHEL meter


def fake_api(session, cache, filt, currency="INR"):
    fx = FX.get(currency, 1.0)
    sku_m = re.search(r"armSkuName eq '([^']+)'", filt)
    ptype_m = re.search(r"priceType eq '([^']+)'", filt)
    sku = sku_m.group(1) if sku_m else None
    ptype = ptype_m.group(1) if ptype_m else None
    if sku not in LINUX_HR:
        return []
    base_hr = LINUX_HR[sku]
    items = []
    if ptype == "Consumption":
        items.append({"productName": "Virtual Machines Series", "meterName": "base", "retailPrice": base_hr * fx})
        items.append({"productName": "Virtual Machines Series Windows", "meterName": "base", "retailPrice": (base_hr + WIN_DELTA_HR) * fx})
        if sku in RHEL_HR:
            items.append({"productName": "Red Hat Enterprise Linux", "meterName": "RHEL", "retailPrice": RHEL_HR[sku] * fx})
    elif ptype == "Reservation":
        items.append({"productName": "Virtual Machines Series", "meterName": "base", "reservationTerm": "1 Year", "retailPrice": base_hr * 730 * 12 * RI1_DISCOUNT * fx})
        items.append({"productName": "Virtual Machines Series", "meterName": "base", "reservationTerm": "3 Years", "retailPrice": base_hr * 730 * 36 * RI3_DISCOUNT * fx})
    return items


@pytest.fixture(autouse=True)
def patch_api(monkeypatch):
    monkeypatch.setattr(C, "_api", fake_api)


def make_workbook(path, rows, sheets=None):
    """sheets: optional list of (title, rows) for multi-sheet tests."""
    wb = Workbook()
    headers = ["Service category", "Service type", "Custom Name", "Region", "Description", "PAYG (Monthly)", "1 Year RI", "3 Year RI"]
    if sheets:
        wb.remove(wb.active)
        for title, sheet_rows in sheets:
            ws = wb.create_sheet(title)
            for i, h in enumerate(headers, 1): ws.cell(1, i, h)
            for r in sheet_rows: ws.append(r)
    else:
        ws = wb.active
        for i, h in enumerate(headers, 1): ws.cell(1, i, h)
        for r in rows: ws.append(r)
    wb.save(path)


def vm_rows(path):
    return list(load_workbook(path)["Virtual Machines"].iter_rows(values_only=True))


# ── C2: multi-sheet workbooks must not silently drop data ──────────────────
def test_all_worksheets_are_parsed(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, None, sheets=[
        ("Group 1", [["Compute", "Virtual Machines", "web-vm", "East US", "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 140.16, 112.13, 84.1]]),
        ("Group 2", [["Compute", "Virtual Machines", "dr-vm", "East US", "1 Standard_D2s_v3 (2 vCPU(s), 8 GB RAM)", 70.08, 56.06, 42.05]]),
    ])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    names = [r[2] for r in rows]
    assert "web-vm" in names, "row from first sheet missing"
    assert "dr-vm" in names, "row from second sheet was silently dropped (multi-sheet bug)"


# ── C3: currency conversion must never silently no-op ──────────────────────
def test_currency_conversion_applies_even_with_only_sql_and_premium_os_rows(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "sql-vm", "East US",
         "1 Standard_E8s_v5 (8 vCPU(s), 64 GB RAM), Red Hat Enterprise Linux, SQL Enterprise (vCPU License)",
         45000.0, 40000.0, 35000.0],
    ])
    C.convert(str(inp), str(out), "INR")
    rows = vm_rows(out)
    compute_payg = rows[2][5]
    # USD raw price would be 367.92; if conversion silently no-ops we'd see
    # that (or something close to it) instead of the ~83x INR-scaled value.
    assert compute_payg > 5000, f"currency conversion appears to have silently no-op'd (got {compute_payg})"


# ── C1: worker exceptions must never vanish silently ────────────────────────
def test_processing_errors_are_caught_and_flagged_not_swallowed(tmp_path, monkeypatch):
    _orig = C.extract_vcpus

    def crasher(desc):
        # D4ds_v5 has no live RHEL meter in fake_api, so it hits the
        # static-fallback path in get_premium_os_license_price(), which is
        # where extract_vcpus() actually gets called.
        if "D4ds_v5" in desc:
            raise RuntimeError("simulated bug")
        return _orig(desc)

    monkeypatch.setattr(C, "extract_vcpus", crasher)

    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "combo-vm", "East US",
         "1 Standard_D4ds_v5 (4 vCPU(s), 16 GB RAM), Red Hat Enterprise Linux, SQL Enterprise (vCPU License)",
         900.0, 700.0, 500.0],
        ["Compute", "Virtual Machines", "web-vm", "East US",
         "3 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM), Windows", 500.0, 400.0, 300.0],
    ])
    # Must not raise out of convert() even though a worker crashes internally.
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    combo_row = next(r for r in rows if r[2] == "combo-vm")
    assert combo_row[8] and "error" in combo_row[8].lower(), "crashed row has no visible remark - failure was silently swallowed"
    web_row = next(r for r in rows if r[2] == "web-vm")
    assert web_row[5] != 500.0, "unrelated row should still have been enriched normally despite the other row's crash"


# ── H2: SKU fallback must never corrupt unrelated SKU families ─────────────
def test_sku_fallback_never_corrupts_m_series():
    variants = C._sku_fallback_variants("Standard_M32ms_v2")
    assert "Standard_M32m_v2" not in variants, "unsafe substring-replace fallback would have produced an invalid SKU"


def test_sku_fallback_still_works_for_legitimate_families():
    assert "Standard_D4_v3" in C._sku_fallback_variants("Standard_D4s_v3")
    assert "Standard_D4d_v5" in C._sku_fallback_variants("Standard_D4ds_v5")


# ── H1: Premium OS license must be its own visible line, not merged silently ─
def test_premium_os_license_gets_its_own_line(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "rhel-vm", "East US",
         "1 Standard_D4ds_v5 (4 vCPU(s), 16 GB RAM), Red Hat Enterprise Linux", 20000.0, 18000.0, 15000.0],
    ])
    C.convert(str(inp), str(out), "INR")
    rows = vm_rows(out)
    descs = [r[4] for r in rows if r[4]]
    assert any("Premium OS License" in d for d in descs), "RHEL license cost was not broken out into its own line"


# ── H4: Azure Hybrid Benefit / BYOL must not be double-charged ─────────────
def test_ahb_byol_suppresses_windows_license_charge(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "byol-vm", "East US",
         "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM), Windows Server (Bring your own license)", 150.0, 140.0, 120.0],
    ])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    descs = [r[4] for r in rows if r[4]]
    assert not any("Windows License" in d for d in descs), "AHB/BYOL VM was still charged a separate Windows license"


# ── C2 (validation): malformed input must raise a clear error, not silently produce an empty file ─
def test_missing_header_raises_clear_error(tmp_path):
    inp = tmp_path / "bad.xlsx"
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    wb.active.append(["not", "a", "calculator", "export"])
    wb.save(inp)
    with pytest.raises(ValueError):
        C.convert(str(inp), str(out), "INR")


def test_unsupported_currency_raises_clear_error(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [["Compute", "Virtual Machines", "x", "East US", "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 100.0, 90.0, 80.0]])
    with pytest.raises(ValueError):
        C.convert(str(inp), str(out), "XYZ")


# ── Totals must be reliable static values that foot exactly, in any viewer ──
def test_totals_are_static_values_and_foot_exactly(tmp_path):
    """
    Totals were briefly implemented as Excel formulas (=SUM(...)), but those
    only display a value once Excel has recalculated and cached it - many
    viewers (including openpyxl/pandas reading the file back without ever
    opening it in Excel) show a formula cell as blank. Totals must be plain
    numbers that already equal the sum of the displayed rows.
    """
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "a", "East US", "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 100.0, 90.0, 80.0],
        ["Compute", "Virtual Machines", "b", "East US", "1 Standard_D2s_v3 (2 vCPU(s), 8 GB RAM)", 50.0, 45.0, 40.0],
    ])
    C.convert(str(inp), str(out), "USD")
    ws = load_workbook(out)["Virtual Machines"]
    total_row = ws.max_row
    total_payg = ws.cell(total_row, 6).value
    assert isinstance(total_payg, (int, float)), f"Total is not a plain number (viewers that don't recalc formulas will show it blank): {total_payg!r}"

    # Foot-check: sum of all displayed PAYG rows above the total must equal the total exactly.
    displayed_sum = 0.0
    for r in range(3, total_row):
        v = ws.cell(r, 6).value
        if isinstance(v, (int, float)):
            displayed_sum += v
    assert round(displayed_sum, 2) == round(total_payg, 2), f"Total ({total_payg}) does not foot to displayed rows ({displayed_sum})"


def test_summary_sheet_totals_are_static_and_match_vm_sheet(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "a", "East US", "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 100.0, 90.0, 80.0],
    ])
    C.convert(str(inp), str(out), "USD")
    wb = load_workbook(out)
    vm_ws = wb["Virtual Machines"]
    vm_total = vm_ws.cell(vm_ws.max_row, 6).value
    summary_ws = wb["Summary"]
    summary_val = summary_ws.cell(3, 3).value
    assert isinstance(summary_val, (int, float)), f"Summary sheet value is not a plain number: {summary_val!r}"
    assert round(summary_val, 2) == round(vm_total, 2), "Summary sheet PAYG doesn't match the Virtual Machines sheet total"


# ── VMSS must be priced through the same pipeline as VMs (own sheet + RI) ──
def test_vmss_gets_its_own_sheet_with_ri_pricing(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual machine scale sets", "web-vmss", "East US",
         "3 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 420.48, 400.0, 380.0],
    ])
    C.convert(str(inp), str(out), "USD")
    wb = load_workbook(out)
    assert "Virtual Machine Scale Sets" in wb.sheetnames, "VMSS was not given its own sheet"
    assert "Others" not in wb.sheetnames or not any(
        "vmss" in str(r[4]).lower() or "scale set" in str(r[1] or "").lower()
        for r in wb["Others"].iter_rows(min_row=3, values_only=True) if r[1] or r[4]
    ), "VMSS row leaked into the 'Others' sheet"
    vmss_ws = wb["Virtual Machine Scale Sets"]
    row = list(vmss_ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    ri1 = row[6]
    payg = row[5]
    assert ri1 != payg, "VMSS 1-Year RI price equals PAYG - RI pricing was not applied (same bug as plain VMs had)"
    assert ri1 < payg, "VMSS RI price should be lower than PAYG"


# ── RI retirement (Azure platform change, 1 Jul 2026) must be explained, not silent ──
def test_retired_ri_series_gets_a_clear_remark_not_silent_payg_fallback(tmp_path, monkeypatch):
    """
    Standard_D2_v3 (Dv3 series) had BOTH 1yr and 3yr Reserved Instance
    purchase/renewal retired by Azure on 2026-07-01. The live API legitimately
    returns no Reservation entries for it any more - this must produce a
    clear, specific remark rather than silently showing RI == PAYG with no
    explanation (which looks like a tool bug).
    """
    def fake_api_no_ri(session, cache, filt, currency="INR"):
        if "priceType eq 'Reservation'" in filt:
            return []
        return fake_api(session, cache, filt, currency)
    monkeypatch.setattr(C, "_api", fake_api_no_ri)

    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [["Compute", "Virtual Machines", "d2v3-vm", "East US", "1 Standard_D2_v3 (2 vCPU(s), 8 GB RAM)", 100.0, 100.0, 100.0]])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    row = next(r for r in rows if r[2] == "d2v3-vm")
    assert row[8] and "retired" in row[8].lower() and "2026" in row[8], f"expected a clear RI-retirement remark, got: {row[8]!r}"


def test_vm_series_classifier_matches_known_retirement_lists():
    cases = [
        ("Standard_D2_v3", "Dv3"), ("Standard_B2s", "Bv1"), ("Standard_F2s", "Fs"),
        ("Standard_F2s_v2", "Fsv2"), ("Standard_F4s", "Fs"), ("Standard_F4s_v2", "Fsv2"),
        ("Standard_D4s_v3", "Dsv3"), ("Standard_D4s_v5", None),
    ]
    for sku, expected in cases:
        assert C._classify_vm_series(sku) == expected, f"{sku}: expected {expected}, got {C._classify_vm_series(sku)}"


# ── NEW: AHB must be scoped independently to Windows OS vs. SQL Server ─────
def test_ahb_on_sql_only_does_not_suppress_windows_license(tmp_path):
    """
    Bug #1 from user report: enabling AHB for SQL was also silently assuming
    Windows AHB, suppressing the real Windows license charge and dumping it
    into the SQL license line via the unaccounted-cost subtraction.
    """
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    # Windows delta = 0.10/hr * 730 = 73.0. SQL AHB means SQL software cost
    # should NOT be charged. orig_payg = compute + windows only (no SQL $).
    compute_usd = LINUX_HR["Standard_D4s_v3"] * 730  # 140.16
    windows_usd = WIN_DELTA_HR * 730  # 73.0
    orig_payg = compute_usd + windows_usd
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "sql-ahb-vm", "East US",
         "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM), Windows, SQL Standard (vCPU License, Bring your own license)",
         orig_payg, orig_payg, orig_payg],
    ])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    # Sub-license line items have a blank Service Category (col A) - only
    # look at those, not the main compute row (whose own Description text
    # legitimately still mentions "SQL" as part of the original label).
    sub_line_descs_vals = {r[4]: r[5] for r in rows if r[4] and not r[0]}
    assert any("Windows License" in d for d in sub_line_descs_vals), "Windows license line missing - AHB-on-SQL incorrectly suppressed Windows charge"
    win_val = next(v for d, v in sub_line_descs_vals.items() if "Windows License" in d)
    assert abs(win_val - windows_usd) < 1.0, f"Windows license value wrong: expected ~{windows_usd}, got {win_val}"
    assert not any("SQL" in d and "License" in d for d in sub_line_descs_vals if d != "Windows License"), \
        "SQL license line should not appear/should be zero - SQL AHB should suppress SQL license charge"


def test_ahb_on_windows_only_does_not_suppress_sql(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    compute_usd = LINUX_HR["Standard_D4s_v3"] * 730
    sql_amount = 200.0
    orig_payg = compute_usd + sql_amount  # no windows delta - AHB covers it
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "win-ahb-vm", "East US",
         "1 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM), Windows Server (Bring your own license), SQL Standard (vCPU License)",
         orig_payg, orig_payg, orig_payg],
    ])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    descs_vals = {r[4]: r[5] for r in rows if r[4]}
    assert not any("Windows License" in d for d in descs_vals), "Windows AHB VM should not show a separate Windows license line"
    assert any("SQL" in d for d in descs_vals), "SQL license line missing - Windows-only AHB incorrectly suppressed SQL charge"


# ── NEW: custom hours must scale PAYG, not always assume 730h ──────────────
def test_custom_hours_scale_payg_not_fixed_730(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "part-time-vm", "East US",
         "1 Standard_D2s_v3 (2 vCPU(s), 8 GB RAM) 400 Hours", 38.4, 30.0, 25.0],
    ])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    row = next(r for r in rows if r[2] == "part-time-vm")
    expected_400h = LINUX_HR["Standard_D2s_v3"] * 400  # 38.4
    expected_730h = LINUX_HR["Standard_D2s_v3"] * 730  # 70.08
    assert abs(row[5] - expected_400h) < 0.5, f"expected ~400h-scaled compute cost {expected_400h}, got {row[5]}"
    assert abs(row[5] - expected_730h) > 5, "compute cost looks like it used the fixed 730h month instead of custom hours"
    assert row[8] and "custom hours" in row[8].lower(), "expected a Remarks note about custom hours being applied"


def test_no_custom_hours_defaults_to_730(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Compute", "Virtual Machines", "full-time-vm", "East US",
         "1 Standard_D2s_v3 (2 vCPU(s), 8 GB RAM)", 70.08, 60.0, 50.0],
    ])
    C.convert(str(inp), str(out), "USD")
    rows = vm_rows(out)
    row = next(r for r in rows if r[2] == "full-time-vm")
    expected_730h = LINUX_HR["Standard_D2s_v3"] * 730
    assert abs(row[5] - expected_730h) < 0.5, f"expected default 730h compute cost {expected_730h}, got {row[5]}"


def test_extract_hours_parses_custom_and_defaults():
    assert C.extract_hours("1 Standard_D2s_v3 (2 vCPU(s), 8 GB RAM) 400 Hours") == 400.0
    assert C.extract_hours("1 Standard_D2s_v3 (2 vCPU(s), 8 GB RAM)") == 730.0
    assert C.extract_hours("1 Standard_D2s_v3, 100 Hours") == 100.0
    assert C.extract_hours("nonsense with 9999 Hours") == 730.0  # out of sane bound -> default


# ── NEW: App Service / AKS / Cosmos DB must not land in "Others" ───────────
def test_app_service_aks_cosmos_are_correctly_classified(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Web", "App Service", "web-app", "East US", "1 P1v3 (2 vCPU(s), 8 GB RAM)", 100.0, 100.0, 100.0],
        ["Containers", "Azure Kubernetes Service (AKS)", "aks-cluster", "East US",
         "3 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 420.48, 420.48, 420.48],
        ["Databases", "Azure Cosmos DB", "cosmos-db", "East US", "400 RU/s", 25.0, 25.0, 25.0],
    ])
    C.convert(str(inp), str(out), "USD")
    wb = load_workbook(out)
    assert "App Service" in wb.sheetnames, "App Service rows should get their own sheet, not fall into Others"
    assert "AKS" in wb.sheetnames, "AKS rows should get their own sheet, not fall into Others"
    assert "Cosmos DB" in wb.sheetnames, "Cosmos DB rows should get their own sheet, not fall into Others"
    if "Others" in wb.sheetnames:
        others_descs = [r[4] for r in wb["Others"].iter_rows(min_row=3, values_only=True) if r[4]]
        assert not any("P1v3" in str(d) or "RU/s" in str(d) for d in others_descs)


# ── NEW: AKS and App Service must get real RI pricing, not PAYG-as-RI ──────
def test_aks_gets_ri_pricing_via_vm_pipeline(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Containers", "Azure Kubernetes Service (AKS)", "aks-cluster", "East US",
         "3 Standard_D4s_v3 (4 vCPU(s), 16 GB RAM)", 420.48, 420.48, 420.48],
    ])
    C.convert(str(inp), str(out), "USD")
    wb = load_workbook(out)
    ws = wb["AKS"]
    row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    payg, ri1 = row[5], row[6]
    assert ri1 != payg, "AKS 1-Year RI price equals PAYG - RI pricing was not applied"
    assert ri1 < payg, "AKS RI price should be lower than PAYG"


def test_app_service_gets_reservation_pricing(tmp_path, monkeypatch):
    def fake_appsvc_api(session, cache, filt, currency="INR"):
        if "Azure App Service" not in filt:
            return fake_api(session, cache, filt, currency)
        fx = FX.get(currency, 1.0)
        if "priceType eq 'Reservation'" not in filt:
            return []
        return [
            {"productName": "App Service", "meterName": "P1v3", "reservationTerm": "1 Year", "retailPrice": 876.0 * fx},
            {"productName": "App Service", "meterName": "P1v3", "reservationTerm": "3 Years", "retailPrice": 1971.0 * fx},
        ]
    monkeypatch.setattr(C, "_api", fake_appsvc_api)

    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Web", "App Service", "web-app", "East US", "1 P1v3 (2 vCPU(s), 8 GB RAM)", 100.0, 100.0, 100.0],
    ])
    C.convert(str(inp), str(out), "USD")
    wb = load_workbook(out)
    ws = wb["App Service"]
    row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    payg, ri1, ri3 = row[5], row[6], row[7]
    assert ri1 != payg, "App Service 1-Year RI price equals PAYG - Reserved App Service Plan pricing was not applied"
    assert abs(ri1 - 876.0 / 12) < 0.5
    assert abs(ri3 - 1971.0 / 36) < 0.5


def test_app_service_non_reservable_tier_flags_remark_instead_of_faking_ri(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    make_workbook(inp, [
        ["Web", "App Service", "basic-app", "East US", "1 B1 (1 vCPU(s), 1.75 GB RAM)", 50.0, 50.0, 50.0],
    ])
    C.convert(str(inp), str(out), "USD")
    wb = load_workbook(out)
    ws = wb["App Service"]
    row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    assert row[8] and "reserv" in row[8].lower(), "Basic tier should be flagged as not reservable, not silently given fake RI pricing"

